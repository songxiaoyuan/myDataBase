#coding: utf-8
# 依赖openpyxl库：http://openpyxl.readthedocs.org/en/latest/
from openpyxl import Workbook
# from openpyxl.compat import range
# from openpyxl.cell import get_column_letter
from openpyxl import load_workbook
import csv
import os
import sys

def xlsx2csv(filename):
    try:
        xlsx_file_reader = load_workbook(filename=filename)
        index = 1
        for sheet in xlsx_file_reader.get_sheet_names():
            # 每个sheet输出到一个csv文件中，文件名用xlsx文件名和sheet名用'_'连接
            print sheet
            csv_filename = sheet+'.csv'
            print csv_filename
            index = index + 1
            # import pdb
            # pdb.set_trace()
            # csv_filename = '{xlsx}_{sheet}.csv'.format(
            #     xlsx=os.path.splitext(filename.replace(' ', '_'))[0],
            #     sheet=sheet.replace(' ', '_'))

            csv_file = file(csv_filename, 'wb')
            csv_file_writer = csv.writer(csv_file)

            sheet_ranges = xlsx_file_reader[sheet]
            for row in sheet_ranges.rows:
                row_container = []
                index =0
                for cell in row:
                    if index >4:
                    	break;
                    index = index +1
                    # if cell.value ==None or cell.value =='' or cell.value ==' ':
                    #     # print 'this is None no data'
                    #     continue
                    if type(cell.value) == unicode:
                        row_container.append(cell.value.encode('utf-8'))
                    else:
                        row_container.append(str(cell.value))
                csv_file_writer.writerow(row_container)
            csv_file.close()

    except Exception as e:
        print(e)

if __name__ == '__main__':
    if len(sys.argv) != 2:
        print('usage: xlsx2csv <xlsx file name>')
    else:
        xlsx2csv(sys.argv[1])
    sys.exit(0)
