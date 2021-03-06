#coding: utf-8
import sys
reload(sys)
sys.setdefaultencoding("utf-8")
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl import load_workbook
import csv
import os
import re

def getTableDic():
	"'这是用来生成所有表的一个字典，键是表的名字，内容就是一个集合，就是所有表头的集合。,读取columns1.txt文件得到"
	tableDic = dict()
	filePath = './columns1.txt'
	file = open(filePath,'rb')
	text = file.readlines()
	# import pdb
	# pdb.set_trace()
	tableName = ''
	for line in text:
		line = line.strip()   #TODO 为什么最后还是有一个\n啊。。。。
		if line =='':
			continue
		if line.startswith('|'):
			# line.strip('|')
			spLine = line.split('|')   #分裂成一个数组，第二个就是字段，但是里面还是有空格的。
			# print spLine
			column = spLine[1]
			column  = column.strip()
			tableDic[tableName].add(column)
		else :
			tableName = line
			if tableName in tableDic:
				print 'the tableName is existed!!!' + tableName
				break
			else:
				tableDic[tableName] = set()
	return tableDic

def readCsvFiles(tableDic_):
	"'这个是用来读取一个文件夹下所有的表的内容,对于每一个csv文件，第2项就是表头，第4项就是表额名字，但是有大写的，应该清除。"
	FolderPath = './csvFiles'
	Files = os.listdir(FolderPath)
	for file in Files:
		file = os.path.join(FolderPath,file)
		csvFile = open(file)
		# import pdb
		# pdb.set_trace()
		tableName = ''
		for line in csvFile:
			if line =='':
				break
			else :
				row = line.split(',')
				column = row[1]
				if tableName == '' and len(row)  == 4:
					tableName = row[3].lower().strip()
				if column =='None':
					print file
					print 'the column is None ' + tableName
					break;
				if tableName in tableDic_:
					if column in tableDic_[tableName]:
						tableDic_[tableName].remove(column)
						#print 'ok'
					else:
						print 'the column is not in tabdic  : ' + column
						print 'the file is  '+file
						print 'the table is '+tableName
	# print tableDic_
	for item in tableDic_:
		print 'the table is : ' +item 
		print tableDic_[item]
def readXlsxFiles(tableDic_):
	"'用来读取某一个文件夹下的所有xlsl文件，然后和数据库生成的字典进行比较"
	FolderPath = './csvFiles'
	Files = os.listdir(FolderPath)
	for file in Files:
		file = os.path.join(FolderPath,file)
		fileReader = load_workbook(file)
		for sheet in fileReader.get_sheet_names():
			tableName = ''
			sheetRanges = fileReader[sheet]
			for row in sheetRanges:
				if row =='':
					break
				else:
					# print row
					# import pdb
					# pdb.set_trace()
					column = row[1].value.strip()
					tableName = row[3].value.strip()
					# line = row.split(',')
					# column = line[1]
					# if tableName == '' and len(line) ==4:
					# 	tableName = line[3].strip()
					# elif tableName == '' and len(line) ==5:
					# 	tableName = line[4].strip()
					# else:
					# 	#do nothing
					# 	a= line[1]
					if column == 'None':
						print file
						print 'the colunm is None ' + tableName
						break
					if column in tableDic_[tableName]:
						tableDic_[tableName].remove(column)
					else:
						print 'the column is not in tabdic : '+column
						print 'the file is '+ file
						print 'the table name is  '+ tableName
	for item in tableDic_:
		if len(tableDic_[item]) == 0 :
			continue
		else:
			print 'the table is  '+item
			print tableDic_[item]

def readDataDicCSVFile(tableDic_):
	filePath = './dataDic.csv'
	csvFile = open(filePath)
	index = 0  
	for line in csvFile:
		if index == 0 :
			index = index+1
			continue
			#确保不读第一行
		else : 
			line = line.split(',')
			# import pdb
			# pdb.set_trace()
			# print line
			column = line[1].strip()
			tableName = ''
			if len(line) ==4 :
				tableName = line[3].lower().strip()
			elif len(line) ==5:
				tableName = line[4].lower().strip()
			else:
				print 'the line is no normal ' + line
			if tableName in tableDic_:
				if column in tableDic_[tableName]:
					tableDic_[tableName].remove(column)
				else:
					print 'the colum :  '+ column + ' is not in :' +tableName
			else:
				print 'the table is not in tabdic: '+ tableName
	for item in tableDic_:
		print 'the table is  ' + item
		print tableDic_[item]

def  writeToOnCSv():
	"'用来读取某一个文件夹下的所有xlsl文件，然后写入一个csv文件"
	FolderPath = './csvFiles'
	csvFileName = 'allToOne.csv'
	csvFile = open(csvFileName,'wb')
	csvFileWriter = csv.writer(csvFile)
	Files = os.listdir(FolderPath)
	index =0
	for file in Files:
		fileName = file.strip('.xlsx')
		print fileName
		file = os.path.join(FolderPath,file)
		fileReader = load_workbook(file)
		for sheet in fileReader.get_sheet_names():
			sheetRanges = fileReader[sheet]
			for row in sheetRanges:
				if row =='':
					break
				else:
					index = index + 1
					rowContainer = []
					for cell in row:
						try:
							cell.value.strip()
							rowContainer.append(cell.value.strip())
						except Exception, e:
							rowContainer.append(' ')
						# import pdb
						# pdb.set_trace()

						# if cell.value.strip():
						# 	pass
						# else:
						# 	print 'the file is '+ file
						# rowContainer.append(cell.value.strip())
					rowContainer.append(fileName)
					csvFileWriter.writerow(rowContainer)
					# column = row[1].value.strip()
					# tableName = row[3].value.strip()
					# if column == 'None':
					# 	print file
					# 	print 'the colunm is None ' + tableName
					# 	break
	print index

def EngToCN():
	"'此函数用来生产中英文对照表"



if __name__ == '__main__':
	# tableDic = getTableDic()
	# print tableDic
	# readCsvFiles(tableDic)
	# readDataDicCSVFile(tableDic)
	# readXlsxFiles(tableDic)
	writeToOnCSv()


